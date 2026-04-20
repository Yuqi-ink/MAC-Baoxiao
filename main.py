import hashlib
import io
import os
import re
import zipfile

import streamlit as st
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader, PdfWriter

# --- 常量 ---
TASK_SAMPLE_LINE = "示例：赴北京参加 XX 学术会议，进行课题调研与交流"
TRANSPORT_OPTIONS = ["飞机", "高铁", "飞机，火车", "其他"]
DUTY_OPTIONS = [
    "博士研究生",
    "硕士研究生",
    "教授",
    "副教授",
    "助理教授",
    "高级工程师",
    "工程师",
]
INV_CAT_OUT = "去程发票"
INV_CAT_HOTEL = "酒店发票"
INV_CAT_RETURN = "返程发票"
INV_CAT_INSURANCE = "保险发票"
INV_CAT_OTHER = "其他支出发票"

# --- 工具函数 ---


def safe_write(sheet, cell_coord, value):
    try:
        sheet[cell_coord] = value
    except AttributeError:
        for range_ in sheet.merged_cells.ranges:
            if cell_coord in range_:
                sheet.cell(row=range_.min_row, column=range_.min_col).value = value
                break


def sanitize_path_component(s: str) -> str:
    if not s:
        return "未填写"
    return re.sub(r'[<>:"/\\|?*]', "_", str(s).strip()) or "未填写"


def travel_fee_bundle_zip_name(d0, person_name: str, dest: str) -> str:
    """总包 ZIP 文件名：x年xx月xx日-姓名-地点差旅费.zip"""
    date_cn = f"{d0.year}年{d0.month:02d}月{d0.day:02d}日"
    n = sanitize_path_component(person_name)
    p = sanitize_path_component(dest)
    return f"{date_cn}-{n}-{p}差旅费.zip"


def invoice_flat_filename(
    start_str: str,
    dest: str,
    person_name: str,
    category_label: str,
    original_name: str,
    *,
    multi_index: int | None = None,
    file_label: str | None = None,
) -> str:
    """发票 ZIP 内扁平文件名；file_label 非空时在类别后追加「-主文件名」（仍保留原扩展名）。"""
    ext = os.path.splitext(original_name)[1]
    if not ext:
        ext = ".dat"
    time_loc = f"{start_str}{sanitize_path_component(dest)}"
    core = f"{time_loc}-{sanitize_path_component(person_name)}-{category_label}"
    label = (file_label or "").strip()
    if label:
        inner = f"{core}-{sanitize_path_component(label)}"
    elif multi_index is not None:
        inner = f"{core}_{multi_index}"
    else:
        inner = core
    return f"{inner}{ext}"


def other_expense_stem_inputs(files, key_prefix: str) -> list[str]:
    """每个其他支出发票一行「归档主文件名」输入，默认原文件名主名。"""
    stems: list[str] = []
    for i, f in enumerate(files):
        digest = hashlib.md5(
            f"{f.name}_{f.size}".encode("utf-8", errors="replace")
        ).hexdigest()[:10]
        stems.append(
            st.text_input(
                f"其他支出「{f.name}」— 压缩包内主文件名（不含后缀，留空则多张时用序号）",
                value=os.path.splitext(f.name)[0],
                key=f"other_stem_{key_prefix}_{i}_{digest}",
            )
        )
    return stems


def image_bytes_to_pdf_bytes(data: bytes) -> bytes:
    src = io.BytesIO(data)
    im = Image.open(src)
    try:
        im = im.convert("RGB")
        try:
            resample = Image.Resampling.LANCZOS
        except AttributeError:
            resample = Image.LANCZOS  # Pillow < 9.1
        im.thumbnail((1600, 1600), resample)
        out = io.BytesIO()
        im.save(out, format="PDF")
        return out.getvalue()
    finally:
        im.close()


def merge_files_to_pdf(
    file_tuples: list[tuple[bytes, str]], out: io.BytesIO
) -> None:
    """将 PNG/JPG 等图片与 PDF 合并为单个 PDF，写入 out（合并后 out 位于末尾，调用方可 getvalue 或 seek(0)）。"""
    writer = PdfWriter()
    readers_keepalive: list[PdfReader] = []
    if not file_tuples:
        writer.add_blank_page(width=595, height=842)
        writer.write(out)
        return
    for data, fname in file_tuples:
        ext = os.path.splitext(fname)[1].lower()
        if ext == ".pdf":
            buf = io.BytesIO(data)
            reader = PdfReader(buf, strict=False)
            readers_keepalive.append(reader)
            for page in reader.pages:
                writer.add_page(page)
        else:
            part = image_bytes_to_pdf_bytes(data)
            buf = io.BytesIO(part)
            reader = PdfReader(buf, strict=False)
            readers_keepalive.append(reader)
            for page in reader.pages:
                writer.add_page(page)
    writer.write(out)


def sync_task_text_area_travelers(travelers: list[tuple[str, str]]) -> None:
    ns = "|".join(f"{n}|{s}" for n, s in travelers)
    header = "\n".join(f"{n} {s}" for n, s in travelers)
    default = f"{header}\n\n{TASK_SAMPLE_LINE}"
    if "_task_ns" not in st.session_state:
        st.session_state._task_ns = ns
        st.session_state.task_area = default
    elif st.session_state._task_ns != ns:
        st.session_state._task_ns = ns
        st.session_state.task_area = default


def build_task_for_excel(task: str, travelers: list[tuple[str, str]]) -> str:
    header = "\n".join(f"{n.strip()} {s.strip()}" for n, s in travelers)
    t = (task or "").strip()
    if t.startswith(header):
        rest = t[len(header) :].lstrip("\n").strip()
        return f"{header}\n\n{rest}" if rest else header
    return f"{header}\n\n{t}" if t else header


def combined_display_name(name1: str, name2: str, has_peer: bool) -> str:
    a = (name1 or "").strip()
    b = (name2 or "").strip()
    if has_peer and b:
        return f"{a}、{b}"
    return a


st.set_page_config(page_title="厦门大学报销自动化助手", layout="wide")
st.title("差旅报销材料自动化整理工具")
st.markdown(
    """
<style>
div[data-testid="stFileUploader"] svg {
    width: 1.35rem !important;
    height: 1.35rem !important;
}
div[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] {
    padding: 0.45rem 0.6rem !important;
    min-height: 2.4rem !important;
}
div[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] span,
div[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] small {
    font-size: 0.92rem !important;
}
.stFileUploader label p {
    font-size: 1.35rem !important;
    font-weight: 600 !important;
    line-height: 1.4 !important;
}
</style>
""",
    unsafe_allow_html=True,
)

st.subheader("1. 基础信息")
row1a, row1b, row1c = st.columns(3)
with row1a:
    name = st.text_input("姓名", value="刘宇琪")
with row1b:
    sid = st.text_input("学号", value="31520240157699")
with row1c:
    duty = st.selectbox("职务", DUTY_OPTIONS, index=0)

has_peer = st.checkbox("有同行人（最多两人同住）", value=False)
name2, sid2 = "", ""
if has_peer:
    pr1, pr2, pr3 = st.columns(3)
    with pr1:
        name2 = st.text_input("同行人姓名", key="peer_name")
    with pr2:
        sid2 = st.text_input("同行人学号", key="peer_sid")
    with pr3:
        st.caption("同行人无需填写职务")

row2a, row2b, row2c = st.columns(3)
with row2a:
    dest = st.text_input("出差地点")
with row2b:
    dates = st.date_input("出差时间范围", [])
same_day_round_trip = len(dates) == 2 and dates[0] == dates[1]
need_hotel_bundle = not same_day_round_trip
with row2c:
    transport_choice = st.selectbox("交通工具", TRANSPORT_OPTIONS)

transport_custom = ""
if transport_choice == "其他":
    transport_custom = st.text_input("请填写具体交通工具")
transport_for_excel = (
    transport_custom.strip() if transport_choice == "其他" else transport_choice
)

st.markdown("**出差事由**")
_travelers: list[tuple[str, str]] = (
    [(name, sid), (name2, sid2)] if has_peer else [(name, sid)]
)
sync_task_text_area_travelers(_travelers)
task = st.text_area(
    "说明（含每人一行「姓名 学号」与事由示例，生成时将按表单姓名学号写入表头）",
    key="task_area",
    height=140,
)

st.subheader("2. 上传附件")
col_inv, col_proof = st.columns(2, gap="large")

with col_inv:
    st.markdown("### 电子发票")
    if has_peer:
        st.caption(
            "双人时请按顺序上传：① 共同发票（酒店、其他支出）→ ② 人员1 交通与保险 → ③ 人员2 交通与保险。"
            + (
                " 当日往返：酒店发票可选（不强制校验）。"
                if same_day_round_trip
                else " 必传：酒店、每人去程/返程；保险可选。"
            )
        )
        st.markdown("#### 共同发票（同住/共用）")
        inv_hotel = st.file_uploader(
            "酒店发票（同住一份，"
            + ("必传" if need_hotel_bundle else "当日往返可选")
            + "；多张时打包自动加序号）",
            type=None,
            accept_multiple_files=True,
            key="inv_hotel_peer",
        )
        inv_other = st.file_uploader(
            "其他支出发票（可选，可多个）",
            type=None,
            accept_multiple_files=True,
            key="inv_other_peer",
        )
        other_inv_list = list(inv_other) if inv_other else []
        other_inv_stems = other_expense_stem_inputs(other_inv_list, "peer")

        st.markdown(f"#### 人员1 · {name or '（请填写姓名）'}")
        inv_out_p1 = st.file_uploader(
            "去程交通工具发票（必传）",
            type=None,
            accept_multiple_files=False,
            key="inv_out_p1_peer",
        )
        inv_return_p1 = st.file_uploader(
            "返程交通工具发票（必传）",
            type=None,
            accept_multiple_files=False,
            key="inv_return_p1_peer",
        )
        inv_insurance_p1 = st.file_uploader(
            "保险发票（可选，可多个）",
            type=None,
            accept_multiple_files=True,
            key="inv_insurance_p1_peer",
        )
        ins_list_p1 = list(inv_insurance_p1) if inv_insurance_p1 else []

        st.markdown(f"#### 人员2 · {name2 or '（请填写同行人姓名）'}")
        inv_out_p2 = st.file_uploader(
            "去程交通工具发票（必传）",
            type=None,
            accept_multiple_files=False,
            key="inv_out_p2_peer",
        )
        inv_return_p2 = st.file_uploader(
            "返程交通工具发票（必传）",
            type=None,
            accept_multiple_files=False,
            key="inv_return_p2_peer",
        )
        inv_insurance_p2 = st.file_uploader(
            "保险发票（可选，可多个）",
            type=None,
            accept_multiple_files=True,
            key="inv_insurance_p2_peer",
        )
        ins_list_p2 = list(inv_insurance_p2) if inv_insurance_p2 else []
    else:
        st.caption(
            "必传：去程交通、返程交通。"
            + (
                " 当日往返：酒店可选（不强制校验）。"
                if same_day_round_trip
                else " 酒店必传。"
            )
            + " 可选：保险、其他支出。"
        )
        inv_out_p1 = st.file_uploader(
            "去程交通工具发票", type=None, accept_multiple_files=False, key="inv_out_solo"
        )
        inv_out_p2 = None
        inv_return_p1 = st.file_uploader(
            "返程交通工具发票", type=None, accept_multiple_files=False, key="inv_return_solo"
        )
        inv_return_p2 = None
        inv_hotel = st.file_uploader(
            "酒店发票（"
            + ("必传，可多个" if need_hotel_bundle else "可选，可多个")
            + "）",
            type=None,
            accept_multiple_files=True,
            key="inv_hotel_solo",
        )
        inv_insurance_p1 = st.file_uploader(
            "保险发票（可选，可多个）",
            type=None,
            accept_multiple_files=True,
            key="inv_insurance_solo",
        )
        inv_insurance_p2 = None
        ins_list_p1 = list(inv_insurance_p1) if inv_insurance_p1 else []
        ins_list_p2 = []
        inv_other = st.file_uploader(
            "其他支出发票（可选，可多个）",
            type=None,
            accept_multiple_files=True,
            key="inv_other_solo",
        )
        other_inv_list = list(inv_other) if inv_other else []
        other_inv_stems = other_expense_stem_inputs(other_inv_list, "solo")

hotel_inv_list = list(inv_hotel) if inv_hotel else []


def transport_skips_traffic_order_proof(choice: str, custom: str) -> bool:
    """含高铁/火车时（含「飞机，火车」联程，或「其他」说明里含高铁/火车），不要求交通类订单/支付截图。"""
    if choice == "其他":
        t = custom or ""
        return "高铁" in t or "火车" in t
    c = choice or ""
    return "高铁" in c or "火车" in c


need_transport_proof = not transport_skips_traffic_order_proof(
    transport_choice, transport_custom
)
has_insurance = bool(ins_list_p1) or bool(ins_list_p2)

with col_proof:
    st.markdown("### 证明材料")
    if has_peer:
        st.caption(
            "双人时请按顺序上传：① 共同材料（酒店订单/支付/水单、保单、其他）→ ② 人员1 交通证明 → ③ 人员2 交通证明。"
            + (
                " 出行含高铁/火车时交通订单/支付证明可不上传。"
                if not need_transport_proof
                else " 交通证明为必传，两人各传一套。"
            )
            + (" 当日往返：酒店类证明可选（不强制校验）。" if same_day_round_trip else "")
        )
        st.markdown("#### 共同证明材料（同住/共用）")
        proof_hotel_order = st.file_uploader(
            "酒店订单截图"
            + ("（当日往返可选）" if same_day_round_trip else ""),
            type=None,
            accept_multiple_files=True,
            key="pho_peer",
        )
        pho_list = list(proof_hotel_order) if proof_hotel_order else []
        proof_hotel_pay = st.file_uploader(
            "酒店支付记录"
            + ("（当日往返可选）" if same_day_round_trip else ""),
            type=None,
            accept_multiple_files=True,
            key="php_peer",
        )
        php_list = list(proof_hotel_pay) if proof_hotel_pay else []
        proof_hotel_bill = st.file_uploader(
            "酒店水单" + ("（当日往返可选）" if same_day_round_trip else ""),
            type=None,
            accept_multiple_files=True,
            key="phb_peer",
        )
        phb_list = list(proof_hotel_bill) if proof_hotel_bill else []
        proof_policy = st.file_uploader(
            "保单首页" + ("（已上传保险发票，请必传）" if has_insurance else "（可选）"),
            type=None,
            accept_multiple_files=True,
            disabled=False,
            key="pp_peer",
        )
        pp_list = list(proof_policy) if proof_policy else []
        proof_other = st.file_uploader(
            "其他证明材料（可选）", type=None, accept_multiple_files=True, key="po_peer"
        )
        po_list = list(proof_other) if proof_other else []

        st.markdown(f"#### 人员1 · {name or '（请填写姓名）'} · 交通")
        proof_transport_p1 = st.file_uploader(
            "往返交通订单、支付记录（必传，出行含高铁/火车除外）",
            type=None,
            accept_multiple_files=True,
            disabled=not need_transport_proof,
            key="pt_p1_peer",
        )
        pt_list_p1 = (
            list(proof_transport_p1) if proof_transport_p1 and need_transport_proof else []
        )

        st.markdown(f"#### 人员2 · {name2 or '（请填写同行人姓名）'} · 交通")
        proof_transport_p2 = st.file_uploader(
            "往返交通订单、支付记录（必传，出行含高铁/火车除外）",
            type=None,
            accept_multiple_files=True,
            disabled=not need_transport_proof,
            key="pt_p2_peer",
        )
        pt_list_p2 = (
            list(proof_transport_p2) if proof_transport_p2 and need_transport_proof else []
        )
    else:
        st.caption(
            (
                "交通类订单/支付证明：出行含高铁/火车时可不上传。"
                if not need_transport_proof
                else "交通类订单/支付证明为必传（往返）。"
            )
            + (" 当日往返：酒店类证明可选（不强制校验）。" if same_day_round_trip else "")
        )
        proof_transport = st.file_uploader(
            "往返交通工具订单截图、支付记录（出行含高铁/火车可不上传）",
            type=None,
            accept_multiple_files=True,
            disabled=not need_transport_proof,
            key="pt_solo",
        )
        proof_transport_p1 = proof_transport
        proof_transport_p2 = None
        pt_list_p1 = list(proof_transport) if proof_transport and need_transport_proof else []
        pt_list_p2 = []

        proof_hotel_order = st.file_uploader(
            "酒店订单截图"
            + ("（当日往返可选）" if same_day_round_trip else ""),
            type=None,
            accept_multiple_files=True,
            key="pho_solo",
        )
        pho_list = list(proof_hotel_order) if proof_hotel_order else []
        proof_hotel_pay = st.file_uploader(
            "酒店支付记录"
            + ("（当日往返可选）" if same_day_round_trip else ""),
            type=None,
            accept_multiple_files=True,
            key="php_solo",
        )
        php_list = list(proof_hotel_pay) if proof_hotel_pay else []
        proof_hotel_bill = st.file_uploader(
            "酒店水单" + ("（当日往返可选）" if same_day_round_trip else ""),
            type=None,
            accept_multiple_files=True,
            key="phb_solo",
        )
        phb_list = list(proof_hotel_bill) if proof_hotel_bill else []
        proof_policy = st.file_uploader(
            "保单首页" + ("（已上传保险发票，请必传）" if has_insurance else "（可选）"),
            type=None,
            accept_multiple_files=True,
            disabled=False,
            key="pp_solo",
        )
        pp_list = list(proof_policy) if proof_policy else []
        proof_other = st.file_uploader(
            "其他证明材料（可选）", type=None, accept_multiple_files=True, key="po_solo"
        )
        po_list = list(proof_other) if proof_other else []


def collect_validation_errors():
    errs = []
    if not (name or "").strip():
        errs.append("请填写姓名")
    if not (sid or "").strip():
        errs.append("请填写学号")
    if not (duty or "").strip():
        errs.append("请填写职务")
    if has_peer:
        if not (name2 or "").strip():
            errs.append("请填写同行人姓名")
        if not (sid2 or "").strip():
            errs.append("请填写同行人学号")
    if not (dest or "").strip():
        errs.append("请填写出差地点")
    if len(dates) != 2:
        errs.append("请选择完整的出差时间范围（开始与结束）")
    if transport_choice == "其他" and not transport_custom.strip():
        errs.append("选择「其他」时请填写具体交通工具")
    if not inv_out_p1:
        errs.append(
            "缺少必传文件：去程交通工具发票（人员1）" if has_peer else "缺少必传文件：去程交通工具发票"
        )
    if has_peer and not inv_out_p2:
        errs.append("缺少必传文件：去程交通工具发票（人员2）")
    if need_hotel_bundle and len(hotel_inv_list) == 0:
        errs.append("缺少必传文件：酒店发票")
    if not inv_return_p1:
        errs.append(
            "缺少必传文件：返程交通工具发票（人员1）" if has_peer else "缺少必传文件：返程交通工具发票"
        )
    if has_peer and not inv_return_p2:
        errs.append("缺少必传文件：返程交通工具发票（人员2）")
    if need_transport_proof:
        if len(pt_list_p1) == 0:
            errs.append(
                "缺少必传文件：往返交通订单/支付记录（人员1）"
                if has_peer
                else "缺少必传文件：往返交通工具订单截图、支付记录"
            )
        if has_peer and len(pt_list_p2) == 0:
            errs.append("缺少必传文件：往返交通订单/支付记录（人员2）")
    if need_hotel_bundle and len(php_list) == 0:
        errs.append("缺少必传文件：酒店支付记录")
    if need_hotel_bundle and len(phb_list) == 0:
        errs.append("缺少必传文件：酒店水单")
    if has_insurance and len(pp_list) == 0:
        errs.append("已上传保险发票，请上传保单首页")
    return errs


if st.button("开始生成报销材料包", use_container_width=True):
    errors = collect_validation_errors()
    if errors:
        for e in errors:
            st.error(e)
        st.stop()

    template_path = os.path.abspath("五定升级版.xlsx")
    if not os.path.isfile(template_path):
        st.error(f"未找到模板文件：{template_path}")
        st.stop()

    _name_tag = sanitize_path_component(combined_display_name(name, name2, has_peer))

    wb = load_workbook(template_path)
    ws = wb.active
    _travelers_excel: list[tuple[str, str]] = (
        [(name, sid), (name2, sid2)] if has_peer else [(name, sid)]
    )
    safe_write(ws, "C6", combined_display_name(name, name2, has_peer))
    safe_write(ws, "E6", duty)
    safe_write(ws, "G6", dest)
    safe_write(ws, "G7", transport_for_excel)
    if len(dates) == 2:
        safe_write(ws, "G5", str(dates[0]))
        safe_write(ws, "H5", str(dates[1]))
    safe_write(ws, "G8", build_task_for_excel(task, _travelers_excel))
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()
    excel_buf.close()

    proof_merge_sequence: list[tuple[bytes, str]] = []
    for f in pt_list_p1:
        proof_merge_sequence.append((f.getvalue(), f.name))
    for f in pt_list_p2:
        proof_merge_sequence.append((f.getvalue(), f.name))
    for f in pho_list:
        proof_merge_sequence.append((f.getvalue(), f.name))
    for f in php_list:
        proof_merge_sequence.append((f.getvalue(), f.name))
    for f in phb_list:
        proof_merge_sequence.append((f.getvalue(), f.name))
    for f in pp_list:
        proof_merge_sequence.append((f.getvalue(), f.name))
    for f in po_list:
        proof_merge_sequence.append((f.getvalue(), f.name))

    merged_pdf_buf = io.BytesIO()
    with st.spinner("正在合并证明材料为 PDF…"):
        merge_files_to_pdf(proof_merge_sequence, merged_pdf_buf)
    merged_pdf_bytes = merged_pdf_buf.getvalue()
    merged_pdf_buf.close()

    start_str = dates[0].isoformat()
    _combo = combined_display_name(name, name2, has_peer)
    inner_zip_entry_name = travel_fee_bundle_zip_name(dates[0], _combo, dest)
    wuding_xlsx_arc = f"厦门大学出差五定审批表_{_name_tag}.xlsx"
    outer_zip_download_name = (
        f"报销材料_{sanitize_path_component(_combo)}_{sanitize_path_component(dest)}_{start_str}.zip"
    )

    inner_zip_buf = io.BytesIO()
    with zipfile.ZipFile(inner_zip_buf, "w", zipfile.ZIP_DEFLATED) as zin:
        zin.writestr(wuding_xlsx_arc, excel_bytes)
        if inv_out_p1:
            arc = invoice_flat_filename(
                start_str, dest, name, INV_CAT_OUT, inv_out_p1.name
            )
            zin.writestr(arc, inv_out_p1.getvalue())
        if has_peer and inv_out_p2:
            arc = invoice_flat_filename(
                start_str, dest, name2, INV_CAT_OUT, inv_out_p2.name
            )
            zin.writestr(arc, inv_out_p2.getvalue())
        for i, inv_hotel_f in enumerate(hotel_inv_list, start=1):
            arc = invoice_flat_filename(
                start_str,
                dest,
                _combo,
                INV_CAT_HOTEL,
                inv_hotel_f.name,
                multi_index=i if len(hotel_inv_list) > 1 else None,
            )
            zin.writestr(arc, inv_hotel_f.getvalue())
        if inv_return_p1:
            arc = invoice_flat_filename(
                start_str, dest, name, INV_CAT_RETURN, inv_return_p1.name
            )
            zin.writestr(arc, inv_return_p1.getvalue())
        if has_peer and inv_return_p2:
            arc = invoice_flat_filename(
                start_str, dest, name2, INV_CAT_RETURN, inv_return_p2.name
            )
            zin.writestr(arc, inv_return_p2.getvalue())
        for i, ins_f in enumerate(ins_list_p1, start=1):
            arc = invoice_flat_filename(
                start_str,
                dest,
                name,
                INV_CAT_INSURANCE,
                ins_f.name,
                multi_index=i if len(ins_list_p1) > 1 else None,
            )
            zin.writestr(arc, ins_f.getvalue())
        for i, ins_f in enumerate(ins_list_p2, start=1):
            arc = invoice_flat_filename(
                start_str,
                dest,
                name2,
                INV_CAT_INSURANCE,
                ins_f.name,
                multi_index=i if len(ins_list_p2) > 1 else None,
            )
            zin.writestr(arc, ins_f.getvalue())
        for i, oth_f in enumerate(other_inv_list, start=1):
            stem_edit = (
                other_inv_stems[i - 1]
                if i <= len(other_inv_stems)
                else os.path.splitext(oth_f.name)[0]
            )
            label = (stem_edit or "").strip() or None
            arc = invoice_flat_filename(
                start_str,
                dest,
                _combo,
                INV_CAT_OTHER,
                oth_f.name,
                multi_index=i if label is None and len(other_inv_list) > 1 else None,
                file_label=label,
            )
            zin.writestr(arc, oth_f.getvalue())
    inner_zip_bytes = inner_zip_buf.getvalue()
    inner_zip_buf.close()

    outer_zip_buf = io.BytesIO()
    with zipfile.ZipFile(outer_zip_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        zout.writestr(inner_zip_entry_name, inner_zip_bytes)
        zout.writestr(wuding_xlsx_arc, excel_bytes)
        zout.writestr("证明材料汇总.pdf", merged_pdf_bytes)
    outer_bytes = outer_zip_buf.getvalue()
    outer_zip_buf.close()

    st.success(
        f"处理完成。总包内含：① {inner_zip_entry_name}（电子发票与五定表 Excel 子压缩包）② {wuding_xlsx_arc}（填好的五定表 Excel）③ 证明材料汇总.pdf"
    )
    st.download_button(
        label=f"下载打包：{outer_zip_download_name}",
        data=outer_bytes,
        file_name=outer_zip_download_name,
        mime="application/zip",
    )
